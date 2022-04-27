from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import copy
import os.path

#----- Manipulate results dictionary ------#
def categorize_slots(input_dict: dict, output_dict: dict):
    #adds layers to the result dictionary to make the MIP-model output more readable
    #fixedSlot          = 1 if slot is fixed, 0 otherwize
    #flexSlot           = 1 if slot is flexible, 0 othervize
    #extSlot            = 1 if slot is extended, 0 otherwize   (adds a possibly extra layer to if slot is fixed)
    #specialty_in_slot  = the index of the specialty assigned to the slot if fixed and -1 of the slot is flexible
    '---parameters---'
    "asdasdasd"
    N                   =   input_dict["N"]
    nDays               =   input_dict["nDays"]
    nRooms              =   input_dict["nRooms"]
    Di                  =   input_dict["Di"]
    Ri                  =   input_dict["Ri"]
    Ci                  =   input_dict["Ci"]
    Gi                  =   input_dict["Gi"]
    RSi                 =   input_dict["RSi"]
    SRi                 =   input_dict["SRi"]
    Si                  =   input_dict["Si"]
    GSi                 =   input_dict["GSi"]
    GRi                 =   input_dict["GRi"]
    gamm                =   output_dict["gamm"]
    lamb                =   output_dict["lamb"]
    x                   =   output_dict["x"]
    
    '---variables---'
    fixed_slot          =   [[0 for _ in range(nDays)] for _ in range(nRooms)]
    flex_slot           =   [[0 for _ in range(nDays)] for _ in range(nRooms)]
    ext_slot            =   [[0 for _ in range(nDays)] for _ in range(nRooms)]
    unass_slot          =   [[0 for _ in range(nDays)] for _ in range(nRooms)]
    specialty_in_slot   =   [[-1 for _ in range(nDays)] for _ in range(nRooms)]
    '---program---'
    for r in Ri:
        for d in Di:
            '---fixed slots---'
            for s in SRi[r]:
                if gamm[s][r][d]>0.5:
                    specialty_in_slot[r][d] = s
                    fixed_slot[r][d]=1
            '---extended slots---'   
            if sum(lamb[s][r][d] for s in Si)>0.5:
                ext_slot[r][d]=1
            '---unassigned slots---'
            if (sum(x[g][r][d][c] for g in Gi for c in Ci)<0.5) and (fixed_slot[r][d]<0.5):
                unass_slot[r][d]=1
            '---flexible slots---'
            if (fixed_slot[r][d]<0.5):
                flex_slot[r][d] = 1
                specialty_in_slot[r][d] = -1         
    '---results---'
    output_dict["fixedSlot"]            =   fixed_slot
    output_dict["flexSlot"]             =   flex_slot
    output_dict["extSlot"]              =   ext_slot
    output_dict["unassSlot"]            =   unass_slot
    output_dict["specialty_in_slot"]    =   specialty_in_slot
    
    
    return output_dict    

def save_results_pre(m):
    statuses=[0,"LOADED","OPTIMAL","INFEASIBLE","INF_OR_UNBD","UNBOUNDED","CUTOFF", "ITERATION_LIMIT",
    "NODE_LIMIT", "TIME_LIMIT", "SOLUTION_LIMIT","INTERUPTED","NUMERIC","SUBOPTIMAL", "USES_OBJ_LIMIT","WORK_LIMIT"]
    result_dict =   {}
    result_dict["status"]=statuses[m.STATUS]
    result_dict["given_more_time"] = False 
    if result_dict["status"] ==  "INFEASIBLE":
        return result_dict
    result_dict["obj"] = m.ObjVal
    result_dict["best_bound"] = m.ObjBound
    result_dict["max_runtime"] = m.Params.TimeLimit
    result_dict["runtime"] = m.Runtime
    result_dict["MIPGap"] = m.MIPGap   
    
    return result_dict

#----- Write to excel ------#
def initiate_excel_book(excel_file_name,input_dict):
        if os.path.exists(excel_file_name):
            return
        wb  = Workbook()
        ws1 = wb.create_sheet("Model_Run",0)
        ws2 = wb.create_sheet("Heuristic_Run",-1)
        ws3 = wb.create_sheet("MSS",-1)
        wb.save(excel_file_name)
    
def write_to_excel_model(file_name,input_dict,output_dict):
    #filename = "myfile.xlsx"
    new_row = []

    try:
        wb = load_workbook(file_name)
        ws = wb.worksheets[0]# select first worksheet
    except FileNotFoundError:
        initiate_excel_book(file_name,input_dict)
        ws = wb.worksheets[0]

    new_row.append(input_dict["nGroups"])
    new_row.append(input_dict["nScenarios"])
    new_row.append(input_dict["F"])
    new_row.append(input_dict["seed"])
    new_row.append(output_dict["obj"])
    new_row.append(output_dict["best_bound"])
    new_row.append(output_dict["MIPGap"])
    new_row.append(output_dict["runtime"])
    new_row.append(output_dict["max_runtime"])
    new_row.append(output_dict["status"])
    new_row.append('-')
    new_row.append('-')
    
    if output_dict["status"] != 0:
        for d in input_dict["Di"]:
            new_row.append(output_dict["bed_occupation"][0][d])
        new_row.append('-')
        new_row.append('-')
        for d in input_dict["Di"]:
            new_row.append(output_dict["bed_occupation"][1][d])

    ws.append(new_row)
    wb.save(file_name)

def write_to_excel_heuristic(excel_file_name,input_dict, best_sol, current_sol, action, global_iter = 0, level = 0, iter = 0, current_gap = -1, only_if_move = False):
    if global_iter > 0 and current_gap > 0:
        new_row = [global_iter, level, iter, best_sol, current_sol, current_gap, str(action)]
    elif global_iter > 0 and current_gap == -1:
        new_row = [global_iter, level, iter, best_sol, current_sol, str(action)]
    elif global_iter == 0 and current_gap > 0:
        new_row = [best_sol, current_sol, current_gap, str(action)]
    else:
        new_row = [best_sol, current_sol, str(action)]
        
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[1]# select first worksheet
    except FileNotFoundError:
        initiate_excel_book(excel_file_name,input_dict)
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[1]
    if only_if_move:
        if action == "MOVE":
            ws.append(new_row)
            wb.save(excel_file_name)
    else:
        ws.append(new_row)
        wb.save(excel_file_name)

def write_to_excel_MSS(excel_file_name,input_dict,output_dict,initial_MSS=False):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[2]
    except FileNotFoundError:
        initiate_excel_book(excel_file_name,input_dict)
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[2]  
    
    if initial_MSS:
        row = ["Initial MSS"]
        ws.append(row) 
    else:
        row = ["Result MSS"]
        ws.append(row) 
        
    """if print_all_cycles:
        cycles = range(1,input_dict["I"]+1)
    else:"""
    row=[]
    cycles = [1]
    for i in cycles:
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        row.append(" ")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = str(d)
            row.append(day)
        ws.append(row)
        row=[]
        for r in input_dict["Ri"]:
            row=[]
            room = str(input_dict["R"][r])
            row.append(room)
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    row.append("-")
                elif output_dict["flexSlot"][r][d] == 1:
                    row.append("#")
                elif (output_dict["unassSlot"][r][d] == 1) and (input_dict["N"][d]>0.5):
                    row.append("?") 
                elif output_dict["fixedSlot"][r][d] == 1:
                    for s in input_dict["Si"]:
                        if output_dict["gamm"][s][r][d] == 1:
                            slotLabel = input_dict["S"][s]
                        if output_dict["lamb"][s][r][d] == 1:
                            slotLabel = slotLabel+"*"
                    row.append(slotLabel)    
            ws.append(row)
    wb.save(excel_file_name)

def write_new_run_header_to_excel(excel_file_name,input_dict,sheet_number=0):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[sheet_number]
    except FileNotFoundError:
        initiate_excel_book(excel_file_name,input_dict)
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[sheet_number]
    
    new_row = []
    new_row.append("------------------- New Run -------------------")
    ws.append(new_row)
    new_row = []
    new_row.append("Date, Time: ")
    now = datetime.now() # current date and time
    date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
    new_row.append(date_time)
    ws.append(new_row)
    
    if sheet_number==0:
        headers_row_model = ['groups','scenarios','flexibility','seed','primal_bound', 'dual_bound', 'gap','runtime','max_runtime','status','-','MC']
        for d in range(1,input_dict["nDays"]+1):
            headers_row_model.append(d)
        headers_row_model.append('-')
        headers_row_model.append('IC')
        for d in range(1,input_dict["nDays"]+1):
            headers_row_model.append(d)
        ws = wb.worksheets[0]
        ws.append(headers_row_model) 
    elif sheet_number==1:
        headers_row_heuristic=["Global iter","Temp level","Temp iter","Best sol","Current sol","Current gap","Action"]
        ws = wb.worksheets[1]
        ws.append(headers_row_heuristic) 
    wb.save(excel_file_name)
    
def write_string_to_excel(excel_file_name, input_dict, string, sheet_number=0):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[sheet_number]
    except FileNotFoundError:
        initiate_excel_book(excel_file_name,input_dict)
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[sheet_number]
    ws.append(string)
    wb.save(excel_file_name)

#----- Print to terminal stats------#
def print_MSS(input_dict, output_dict, print_all_cycles = False):

    if print_all_cycles:
        cycles = range(1,input_dict["I"]+1)
    else:
        cycles = range(1,2)
        
    print("Planning period modified MSS")
    print("-----------------------------")
    for i in cycles:
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for r in input_dict["Ri"]:
            room = "{0:>8}".format(input_dict["R"][r]+"|")
            print(room, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    print("{0:<5}".format("-"), end="")
                elif (output_dict["unassSlot"][r][d] == 1):
                    print("{0:<5}".format("?"), end="") 
                elif output_dict["flexSlot"][r][d] == 1:
                    print("{0:<5}".format("#"), end="") 
                elif output_dict["fixedSlot"][r][d] == 1:
                    if output_dict["specialty_in_slot"][r][d]>=0:
                        slotLabel = input_dict["S"][output_dict["specialty_in_slot"][r][d]]
                    if output_dict["extSlot"][r][d] == 1:
                        slotLabel = slotLabel+"*"
                    """output_dict["specialty_in_slot"][s][r]
                    if output_dict["gamm"][s][r][d] == 1:
                        slotLabel = input_dict["S"][s]
                    if output_dict["lamb"][s][r][d] == 1:
                        slotLabel = slotLabel+"*"""
                    print("{0:<5}".format(slotLabel), end="")
            print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()

def print_expected_minutes(input_dict, output_dict):

    print("Expected number of planned operation minutes per slot")
    print("-----------------------------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for r in input_dict["Ri"]:
            room = "{0:>8}".format(input_dict["R"][r]+"|")
            print(room, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    print("{0:<5}".format("-"), end="")
                else:
                    operations = 0
                    for g in input_dict["Gi"]: 
                        for c in input_dict["Ci"]:
                            operations += (input_dict["Pi"][c])*(output_dict["x"][g][r][d][c]*(input_dict["L"][g]+input_dict["TC"]))
                    if operations > 0:
                        operations_str = "{:.0f}".format(operations)
                    else:
                        operations_str = "{:.0f}".format(operations)
                    print("{0:<5}".format(str(operations_str)), end="")
            print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        print()  

def print_expected_bed_util_percent(input_dict, output_dict):
    print("Expected bed ward utilization")
    print("-----------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for w in input_dict["Wi"]:
            ward = "{0:>8}".format(input_dict["W"][w]+"|")
            print(ward, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                total = output_dict["bed_occupation"][w][d]/input_dict["B"][w][d]
                total = "{:.2f}".format(total)
                print("{0:<5}".format(str(total)), end="")
            print()
                
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()

def print_expected_bed_util(input_dict, output_dict):
    
    print("Expected bed ward utilization")
    print("-----------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for w in input_dict["Wi"]:
            ward = "{0:>8}".format(input_dict["W"][w]+"|")
            print(ward, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                total = output_dict["bed_occupation"][w][d]
                total = "{:.1f}".format(total)
                print("{0:<5}".format(str(total)), end="")
            print()
                
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()

def print_expected_que(input_dict, output_dict):
    share_met_demand    =   [0 for _ in range(input_dict["nGroups"])]
    minutes_met_demand  =   [0 for _ in range(input_dict["nGroups"])]    
    
    print("           E[unmet #operations]  E[fracton operated]  E[que] ")
    print("        -------------------------------------------------------")
    for g in input_dict["Gi"]:
        gg              =   "{0:<2}".format(str(g))
        name            =   "{0:<4}".format(str(input_dict["G"][g]))
        print("%s %s|        " %(gg,name), end= " ")
        que                     =   sum(input_dict["Pi"][c] * output_dict["a"][g][c] for c in input_dict["Ci"])
        que                     =   "{:.2f}".format(que)
        que                     =   "{0:>7}".format(str(que))
        print("%s            " %que, end= " ")
        share_met_demand[g]     =   0
        minutes_met_demand[g]   =   0
        
        for c in input_dict["Ci"]: 
            if input_dict["Q"][g][c]>0:
                share_met_demand[g]     +=  input_dict["Pi"][c] * sum(output_dict["x"][g][r][d][c] for r in input_dict["Ri"] for d in input_dict["Di"])/input_dict["Q"][g][c]
            else:
                share_met_demand[g]     +=  input_dict["Pi"][c] * 1
            minutes_met_demand[g]   +=  input_dict["Pi"][c] * output_dict["a"][g][c]*input_dict["L"][g]
            
            if sum(output_dict["x"][g][r][d][c] for r in input_dict["Ri"] for d in input_dict["Di"] for c in input_dict["Ci"] ) < 0.5:
                share_met_demand[g] = 0
        share               =   "{:.3f}".format(share_met_demand[g])    
        share               =   "{0:>6}".format(share)
        min                 =   "{:.1f}".format(minutes_met_demand[g])
        min                 =   "{0:>9}".format(min)
        print("%s        %s" %(share,min) )
    print()
    print()

def print_solution_performance(input, results):
    print()
    print("Solution performance")
    print("--------------------------------")
    print("Status:          ", end="")
    print(results["status"])
    print("Runtime:         ", end="")
    runtime = "{:.0f}".format(results["runtime"])
    print("{0:>10}".format(str(runtime)+"s"))
    print("Objective :      ", end="")
    obj = "{:.1f}".format(results["obj"])
    print("{0:>10}".format(str(obj)))
    print("Dual bound:      ", end="")
    dual = "{:.1f}".format(results["best_bound"])
    print("{0:>10}".format(str(dual)))
    print("Optimality gap:  ", end="")
    gap = "{:.1f}".format(results["MIPGap"]*100)
    print("{0:>10}".format(str(gap)+"%"))
    print()
    
def print_heuristic_iteration_header(mip = True, gap = True):
    if mip:
        print("{0:<15}".format("Global iter"), end="")
        print("{0:<15}".format("Temp level"), end="")
        print("{0:<15}".format("Temp iter"),end="")
    print("{0:<15}".format("Best sol"),end="")
    print("{0:<15}".format("Current sol"),end="")
    if gap:
        print("{0:<15}".format("Current gap"),end="")
    print("{0:<15}".format("Swap type"),end="")
    print("{0:<15}".format("Action"),end="")
    print("{0:<15}".format("Time"),end="")
    print()

def print_heuristic_iteration(best_sol, current_sol, swap_type, action, time, global_iter = 0, level = 0, levels = [0], iter = 0, level_iters = [0], current_gap = -1):
    if global_iter != 0:
        level_str = str(level)+"/"+str(len(levels))
        iter_str = str(iter)+"/"+str(level_iters[level-1])
        print("{0:<15}".format(global_iter), end="")
        print("{0:<15}".format(level_str), end="")
        print("{0:<15}".format(iter_str),end="")
    print("{0:<15}".format("{:.1f}".format(best_sol)),end="")
    print("{0:<15}".format("{:.1f}".format(current_sol)),end="")
    if current_gap > -1:
        print("{0:<15}".format(str("{:.1f}".format(current_gap*100))+"%"),end="")
    print("{0:<15}".format(swap_type),end="")
    print("{0:<15}".format(action),end="")
    time_str = "{:.0f}".format(time)+'s'
    print("{0:<15}".format(time_str),end="")
    print()

