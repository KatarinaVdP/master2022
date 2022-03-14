from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


def categorize_slots(input_dict, output_dict):
    
    fixed_slots = [[0 for _ in range(input_dict["nDays"])] for _ in range(input_dict["nRooms"])]
    flex_slot = [[0 for _ in range(input_dict["nDays"])] for _ in range(input_dict["nRooms"])]
    ext_slot = [[0 for _ in range(input_dict["nDays"])] for _ in range(input_dict["nRooms"])]
    
    days_in_cycle = int(input_dict["nDays"]/input_dict["I"])
    flex_count = 0
    fixed_count = 0
    
    for r in input_dict["Ri"]:
        day_in_cycle=0
        for d in range(input_dict["nDays"]):
            if day_in_cycle >= days_in_cycle:
                    day_in_cycle=0
            if sum(output_dict["delt"][s][r][d][c] for s in input_dict["Si"] for c in input_dict["Ci"])>0.5:
                flex_slot[r][d] = 1
                flex_count += 1
                for dd in range(input_dict["nDays"]):
                        if (dd % days_in_cycle) == day_in_cycle:
                            flex_slot[r][dd]=1
            if sum(output_dict["gamm"][s][r][d] for s in input_dict["Si"])>0.5:
                fixed_slots[r][d] = 1
                fixed_count += 1
                if sum(output_dict["lamb"][s][r][d] for s in input_dict["Si"])>0.5:
                    ext_slot[r][d]=1
            day_in_cycle += 1
    
    output_dict["fixedSlot"]    = fixed_slots
    output_dict["flexSlot"]     = flex_slot
    output_dict["extSlot"]      = ext_slot
    return output_dict    

def print_MSS(input_dict, output_dict, print_all_cycles = False):

    if print_all_cycles:
        cycles = range(1,input_dict["I"]+1)
    else:
        cycles = range(1,2)
        
    print("Planning period modified MSS")
    print("-----------------------------")
    for i in cycles:
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for r in input_dict["Ri"]:
            room = "{0:>8}".format(input_dict["R"][r]+"|")
            print(room, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    print("{0:<5}".format("-"), end="")
                elif output_dict["flexSlot"][r][d] == 1:
                    print("{0:<5}".format("#"), end="") 
                elif (output_dict["fixedSlot"][r][d] == 0) & (output_dict["flexSlot"][r][d] == 0):
                    print("{0:<5}".format("?"), end="") 
                elif output_dict["fixedSlot"][r][d] == 1:
                    for s in input_dict["Si"]:
                        if output_dict["gamm"][s][r][d] == 1:
                            slotLabel = input_dict["S"][s]
                        if output_dict["lamb"][s][r][d] == 1:
                            slotLabel = slotLabel+"*"
                    print("{0:<5}".format(slotLabel), end="")
            print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()

def print_expected_operations(input_dict, output_dict):

    print("Expected number of planned operations per slot")
    print("-----------------------------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for r in input_dict["Ri"]:
            room = "{0:>8}".format(input_dict["R"][r]+"|")
            print(room, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    print("{0:<5}".format("-"), end="")
                else:
                    operations = 0
                    for g in input_dict["Gi"]: 
                        for c in input_dict["Ci"]:
                            operations += (input_dict["Pi"][c])*(output_dict["x"][g][r][d][c])
                    if operations > 0:
                        operations_str = "{:.1f}".format(operations)
                    else:
                        operations_str = "{:.0f}".format(operations)
                    print("{0:<5}".format(str(operations_str)), end="")
            print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        print()
        print()
        print()
        
def print_expected_minutes(input_dict, output_dict):

    print("Expected number of planned operation minutes per slot")
    print("-----------------------------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for r in input_dict["Ri"]:
            room = "{0:>8}".format(input_dict["R"][r]+"|")
            print(room, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                if input_dict["N"][d] == 0:
                    print("{0:<5}".format("-"), end="")
                else:
                    operations = 0
                    for g in input_dict["Gi"]: 
                        for c in input_dict["Ci"]:
                            operations += (input_dict["Pi"][c])*(output_dict["x"][g][r][d][c]*(input_dict["L"][g]+input_dict["TC"]))
                    if operations > 0:
                        operations_str = "{:.0f}".format(operations)
                    else:
                        operations_str = "{:.0f}".format(operations)
                    print("{0:<5}".format(str(operations_str)), end="")
            print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        print()
        print()
        print()
        
def bed_occupation(input_dict, output_dict, w, d, c):
    occupation = 0
    for g in input_dict["GWi"][w]:
        for r in input_dict["Ri"]:
            for dd in range(max(0,d+1-input_dict["J"][w]), d+1):
                occupation += input_dict["P"][w][g][d-dd] * output_dict["x"][g][r][dd][c] + output_dict["v"][w][d]
    return occupation
        
def print_expected_bed_util(input_dict, output_dict):

    """bed_occupation =[[[0 for _ in range(input_dict["nScenarios"])] for _ in range(input_dict["nDays"])] for _ in range(input_dict["nWards"])]
    
    for w in input_dict["Wi"]:
        for d in input_dict["Di"]:
            for c in input_dict["Ci"]:
                bed_occupation[w][d][c]= sum(input_dict["P"][w][g][d-dd] * output_dict["x"][g][r][dd][c] for g in input_dict["GWi"][w] for r in input_dict["Ri"] for dd in range(max(0,d+1-input_dict["J"][w]),d+1)) + input_dict["Y"][w][d]"""
    
    print("Expected bed ward utilization")
    print("-----------------------------")
    for i in range(1,input_dict["I"]+1):
        print("Cycle: ", i)
        print("        ", end="")
        nDaysInCycle = int(input_dict["nDays"]/input_dict["I"])
        firstDayInCycle = int(nDaysInCycle*(i-1)+1)
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            day = "{0:<5}".format(str(d))
            print(day, end="")
        print()
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        for w in input_dict["Wi"]:
            ward = "{0:>8}".format(input_dict["W"][w]+"|")
            print(ward, end="")
            for d in range(firstDayInCycle-1,firstDayInCycle+nDaysInCycle-1):
                total = output_dict["bed_occupation"][w][d]
                total = "{:.1f}".format(total)
                print("{0:<5}".format(str(total)), end="")
            print()
                
        print("        ", end="")
        for d in range(firstDayInCycle,firstDayInCycle+nDaysInCycle):
            print("-----",end="")
        print()
        
def print_operations_per_group(input_dict, output_dict):
    
    print("Total number of operations per surgery group across all scenarios")
    print("------------------------------------------------------------------")
    
    for g in input_dict["Gi"]:
        group = "{0:<5}".format(input_dict["G"][g])+"|"
        print(group, end="")
        operations = 0
        operations = sum(output_dict["x"][g][r][d][c] for c in input_dict["Ci"] for d in input_dict["Di"] for r in input_dict["Ri"])
        print("{0:>5}".format(operations))
        
    print()
    print()
    print()
    print()

def print_que(input_dict, output_dict):
    share_met_demand    =   [0 for _ in range(input_dict["nGroups"])]
    minutes_met_demand  =   [0 for _ in range(input_dict["nGroups"])]    
    
    print("           E[unmet #operations]  E[fracton operated]  E[que] ")
    print("        -------------------------------------------------------")
    for g in input_dict["Gi"]:
        gg              =   "{0:<2}".format(str(g))
        name            =   "{0:<4}".format(str(input_dict["G"][g]))
        print("%s %s|        " %(gg,name), end= " ")
        que                     =   sum(input_dict["Pi"][c] * output_dict["a"][g][c] for c in input_dict["Ci"])
        que                     =   "{:.2f}".format(que)
        que                     =   "{0:>7}".format(str(que))
        print("%s            " %que, end= " ")
        share_met_demand[g]     =   0
        minutes_met_demand[g]   =   0
        
        for c in input_dict["Ci"]: 
            if input_dict["Q"][g][c]>0:
                share_met_demand[g]     +=  input_dict["Pi"][c] * sum(output_dict["x"][g][r][d][c] for r in input_dict["Ri"] for d in input_dict["Di"])/input_dict["Q"][g][c]
            else:
                share_met_demand[g]     +=  input_dict["Pi"][c] * 1
            minutes_met_demand[g]   +=  input_dict["Pi"][c] * output_dict["a"][g][c]*input_dict["L"][g]
            
            if sum(output_dict["x"][g][r][d][c] for r in input_dict["Ri"] for d in input_dict["Di"] for c in input_dict["Ci"] ) < 0.5:
                share_met_demand[g] = 0
        share               =   "{:.3f}".format(share_met_demand[g])    
        share               =   "{0:>6}".format(share)
        min                 =   "{:.1f}".format(minutes_met_demand[g])
        min                 =   "{0:>9}".format(min)
        print("%s        %s" %(share,min) )
    print()
    print()

def write_to_excel(file_name,input_dict,output_dict):
    #filename = "myfile.xlsx"
    new_row = []

    try:
        wb = load_workbook(file_name)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        headers_row = ['groups','scenarios','flexibility','seed','primal_bound', 'dual_bound', 'gap','runtime','max_runtime','status','-','MC',]
        for d in range(1,input_dict["nDays"]+1):
            headers_row.append(d)
        headers_row.append('-')
        headers_row.append('IC')
        for d in range(1,input_dict["nDays"]+1):
            headers_row.append(d)
        wb = Workbook()
        ws = wb.active
        ws.append(headers_row)

    new_row.append(input_dict["number_of_groups"])
    new_row.append(input_dict["nScenarios"])
    new_row.append(input_dict["F"])
    new_row.append(input_dict["seed"])
    new_row.append(output_dict["obj"])
    new_row.append(output_dict["best_bound"])
    new_row.append(output_dict["MIPGap"])
    new_row.append(output_dict["runtime"])
    new_row.append(output_dict["max_runtime"])
    new_row.append(output_dict["status"])
    new_row.append('-')
    new_row.append('-')
    
    if output_dict["status"] != 0:
        for d in input_dict["Di"]:
            new_row.append(output_dict["bed_occupation"][0][d])
        new_row.append('-')
        new_row.append('-')
        for d in input_dict["Di"]:
            new_row.append(output_dict["bed_occupation"][1][d])

    ws.append(new_row)
    wb.save(file_name)
    
def write_header_to_excel(file_name, type_of_header: str):
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]
    new_row = []
    
    if type_of_header == "begin_new":
        new_row.append("------------------- New Run -------------------")
        ws.append(new_row)
        wb.save(file_name)
        new_row = []
        new_row.append("new")
        new_row.append("heuristic")
        new_row.append("run")
        new_row.append("")
        new_row.append("Date:")
        now = datetime.now() # current date and time
        date_time = now.strftime("%m/%d/%Y, %H:%M:%S")
        new_row.append(date_time)
        ws.append(new_row)
        wb.save(file_name)
    elif type_of_header == "first_iteration":
        new_row.append("------------------- Begin Heuristic -------------------")
        ws.append(new_row)
        wb.save(file_name)
    else:
        print('Invalid input in write_header_to_excel()')
        return
    
def print_solution_performance(input, results):
    print("Solution performance")
    print("--------------------------------")
    print("Status:          ", end="")
    print(results["status"])
    print("Runtime:         ", end="")
    runtime = "{:.0f}".format(results["runtime"])
    print("{0:>10}".format(str(runtime)+"s"))
    print("Objective :      ", end="")
    obj = "{:.1f}".format(results["obj"])
    print("{0:>10}".format(str(obj)))
    print("Dual bound:      ", end="")
    dual = "{:.1f}".format(results["best_bound"])
    print("{0:>10}".format(str(dual)))
    print("Optimality gap:  ", end="")
    gap = "{:.1f}".format(results["MIPGap"]*100)
    print("{0:>10}".format(str(gap)+"%"))
    print()
    
def print_heuristic_iteration_header():
    print("{0:<15}".format("Global iter"), end="")
    print("{0:<15}".format("Temp level"), end="")
    print("{0:<15}".format("Temp iter"),end="")
    print("{0:<15}".format("Best sol"),end="")
    print("{0:<15}".format("Current sol"),end="")
    print("{0:<15}".format("Current gap"),end="")
    print("{0:<15}".format("Action"),end="")
    print()

def print_heuristic_iteration(global_iter, level, levels, iter, level_iters, best_sol, current_sol, current_gap, action):
    level_str = str(level)+"/"+str(len(levels))
    iter_str = str(iter)+"/"+str(level_iters[level-1])
    print("{0:<15}".format(global_iter), end="")
    print("{0:<15}".format(level_str), end="")
    print("{0:<15}".format(iter_str),end="")
    print("{0:<15}".format("{:.1f}".format(best_sol)),end="")
    print("{0:<15}".format("{:.1f}".format(current_sol)),end="")
    print("{0:<15}".format(str("{:.1f}".format(current_gap*100))+"%"),end="")
    print("{0:<15}".format(action),end="")
    print()
    