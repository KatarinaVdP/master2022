from functions_input import *
from functions_output import *
from model_mip import *
from heuristic_greedy_construction import *
from openpyxl import Workbook
from openpyxl import load_workbook


def write_to_excel_problem_size(excel_file_name: str, results_mip: dict, flex: int, nScenarios: int, seed: int, bed_cap_factor: float, nGroups:int):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("problem_size",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("nGroups")
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("bed_cap_factor")
        header_row.append("primal")
        header_row.append("dual")
        header_row.append("MIPGap")
        header_row.append("runtime")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(nGroups)
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(bed_cap_factor)
    new_row.append(results_mip["obj"])
    new_row.append(results_mip["best_bound"])
    new_row.append(results_mip["MIPGap"])
    new_row.append(results_mip["runtime"])
    ws.append(new_row)
    wb.save(excel_file_name)  

num_groups                  =   [9]
num_scenarios               =   [50]
bed_caps                    =   [0.6]
flexibilities               =   [0.1]

time_to_mip                 =   3600
seeds                       =   [1,2,3,4,5]
excel_file_name             =   'input_output/bed_ward_complexity06.xlsx'

for cap in bed_caps:
    for ng in num_groups:
        input_file_name         =   choose_correct_input_file(ng)
        input                   =   read_input(input_file_name)
        if ng ==25:
            input           =   change_ward_capacity(input, "MC",72.4*cap,56*cap)
            input           =   change_ward_capacity(input, "IC",14.5*cap,6.1*cap) 
        elif ng ==9:
            input           =   change_ward_capacity(input, "MC",60*cap,49*cap)
            input           =   change_ward_capacity(input, "IC",11*cap,6*cap)  
        elif ng ==5:
            input           =   change_ward_capacity(input, "MC",50.5*cap,42*cap)
            input           =   change_ward_capacity(input, "IC",9.1*cap,5.6*cap)
        else:
            print('No scaling factors exists')
            break
        for seed in seeds:
            for ns in num_scenarios:
                input               =   generate_scenarios(input, ns, seed)
                for flex in flexibilities:
                    results, input          =   run_model_mip(input,flex,time_to_mip,expected_value_solution=False,print_optimizer = True)
                    write_to_excel_problem_size(excel_file_name, results,flex,ns,seed,cap,ng)
                    print("nGroups: %i  nScenarios: %i  flex: %.2f  bed_cap_factor: %.2f  primal: %.1f  dual: %.1f MIPgap: %.3f runtime: %.1f "%(ng,ns,flex,cap, results["obj"], results["best_bound"], results["MIPGap"],results["runtime"]))
            
