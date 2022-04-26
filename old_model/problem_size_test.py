from functions_input import *
from functions_output import *
from model_mip import *
from heuristic_greedy_construction import *
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import copy
import os.path


def print_heuristic_vs_fixed(results_mip: dict, results_heuristic: dict, flex: int, nScenarios: int, seed: int,max_time_fixed_mip: int):
    
    mip_str_p = "mip_p: "+ str("{:.1f}".format(results_mip["obj"]))
    mip_str_d = "mip_d: "+ str("{:.1f}".format(results_mip["best_bound"]))
    heur_str =  "heur: "  + str("{:.1f}".format(results_heuristic["obj"]))
    preformance_str = "diff: " + str("{:.2f}".format((results_mip["obj"]-results_heuristic["obj"])/results_mip["obj"])) 
    

    print("{0:<15}".format(flex), end="")
    print("{0:<15}".format(nScenarios), end="")
    print("{0:<15}".format(seed), end="")
    print("{0:<15}".format("mip_time: " + str(max_time_fixed_mip)), end="")
    print("{0:<15}".format(mip_str_p), end="")
    print("{0:<15}".format(mip_str_d),end="")
    print("{0:<15}".format(heur_str),end="")
    print("{0:<15}".format(preformance_str),end="")
    print() 

def write_to_excel_problem_size(excel_file_name: str, results_mip: dict, flex: int, nScenarios: int, seed: int, bed_cap_factor: float):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("problem_size",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("bed_cap_factor")
        header_row.append("primal")
        header_row.append("dual")
        header_row.append("MIPGap")
        header_row.append("runtime")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(bed_cap_factor)
    new_row.append(results_mip["obj"])
    new_row.append(results_mip["best_bound"])
    new_row.append(results_mip["MIPGap"])
    new_row.append(results_mip["runtime"])
    ws.append(new_row)
    wb.save(excel_file_name)  

num_groups                  =   [ 9, 25]
num_scenarios               =   [10,50,100,200]
bed_caps                    =   [1, 0.5]
flexibilities               =   [0.1]

time_to_mip                 =   20
seed                        =   1
excel_file_name             =   'input_output/problem_size.xlsx'

for ng in num_groups:
    input_file_name         =   choose_correct_input_file(ng)
    input                   =   read_input(input_file_name)
    if ng ==   25:
        change_demand(input, 1.35, print_minutes = False)
    for ns in num_scenarios:
        input               =   generate_scenarios(input, ns, seed)
        for flex in flexibilities:
            for cap in bed_caps:
                input                   =   change_ward_capacity(input, "MC", 60*cap,  49*cap)
                input                   =   change_ward_capacity(input, "IC", 11*cap,  7*cap)
                results, input          =   run_model_mip(input,flex,time_to_mip,expected_value_solution=False,print_optimizer = False)
                write_to_excel_problem_size(excel_file_name, results,flex,ns,seed,cap)
                print("nGroups: %i  nScenarios: %i  flex: %.2f  bed_cap_factor: %.2f  primal: %.1f  dual: %.1f MIPgap: %.3f runtime: %.1f "%(ng,ns,flex,cap, results["obj"], results["best_bound"], results["MIPGap"],results["runtime"]))
        
