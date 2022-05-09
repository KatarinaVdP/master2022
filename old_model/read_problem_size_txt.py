from openpyxl import Workbook
from openpyxl import load_workbook


def write_to_excel_problem_size(excel_file_name: str, r:int,c:int,i:int,p:int,b:int,p_r:int,p_c:int,p_i:int,p_p:int,p_b:int):
    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("problem_size",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("rows")
        header_row.append("cols")
        header_row.append("integer")
        header_row.append("binary")
        header_row.append("nonzero")
        header_row.append("")
        header_row.append("rows")
        header_row.append("cols")
        header_row.append("integer")
        header_row.append("binary")
        header_row.append("nonzero")

        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(r)
    new_row.append(c)
    new_row.append(i)
    new_row.append(b)
    new_row.append(p)
    new_row.append("")
    new_row.append(p_r)
    new_row.append(p_c)
    new_row.append(p_i)
    new_row.append(p_b)
    new_row.append(p_p)
    ws.append(new_row)
    wb.save(excel_file_name)  

excel_file_name             =   'input_output/problem_size_row_col_new.xlsx'
txt_file_name               =   'input_output/saved_results/test.txt'

with open(txt_file_name ) as f:
    lines = f.readlines()
    count=0
    for line in lines:
        l=line.split()
        if len(l)>=7:
            if l[0] == "Optimize" and l[1]== "a":
                print("hei")
                r=int(l[4])
                c=int(l[6])
                p=int(l[9])
                lineVar= lines[count+2]
                l_v=lineVar.split()
                i=int(l_v[4])
                b_pre=l_v[6].split("(")
                b=int(b_pre[1])
                """print("prob")
                print(r)
                print(c)
                print(i)
                print(b)"""
            if l[0] == "Presolved:":    
                print("på deg")
                p_r=int(l[1])
                p_c=int(l[3])
                p_p=int(l[5])
                lineVar= lines[count+1]
                l_v=lineVar.split()
                p_i=int(l_v[4])
                b_pre=l_v[6].split("(")
                p_b=int(b_pre[1])
                """ print("pre")
                print(p_r)
                print(p_c)
                print(p_i)
                print(p_b)"""
                write_to_excel_problem_size(excel_file_name,r,c,i,p,b,p_r,p_c,p_i,p_p,p_b)
        count+=1
        print(count)