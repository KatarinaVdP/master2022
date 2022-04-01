from functions_input import *
from functions_output import *
#from model_cutting_stock import *
from model_mip import *
from heuristic_greedy_construction import *
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import copy
import os.path

def write_to_excel(excel_file_name: str, results_heuristic: dict, flex: int, nScenarios: int, seed: int):
    #filename = "myfile.xlsx"

    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]# select first worksheet
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("stability analysis",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("heuristic_sol")
        header_row.append("heuristic_time")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(results_heuristic["obj"])
    new_row.append(results_heuristic["heuristic_time"])
    ws.append(new_row)
    wb.save(excel_file_name)  

number_of_groups            =   25
nScenarios_initial_sol      =   1
time_to_mip                 =   20
mip_seeds                   =   [i for i in range(100,131)]
file_name                   =   choose_correct_input_file(number_of_groups)
excel_file_name             =   'input_output/stability_test.xlsx'

flexibilities               =   [0]
num_sols_to_investigate     =   5
#nScens                      =   [10,20]
#seeds                       =   [i for i in range(1,3)]
nScens                      =   [10,20,30,40,50,60,70,80,90,100,
                                110,120,130,140,150,160,170,180,190,200, 
                                210,220,230,240,250,260,270,280,290,300]
seeds                       =   [i for i in range(1,31)]
book = Workbook()
sheet = book.active
current_row_outlay=2
col_index = 0
alphabeth= ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M', 'N', 'O', 'P', 'Q','R', 'S','T','U','V','W','X','Y','Z']
input                   =   read_input(file_name)
for flex in flexibilities:
    for iter in range(num_sols_to_investigate):
        input                   =   generate_scenarios(input, nScenarios_initial_sol, mip_seeds[iter])
        results, input          =   run_model_mip(input,flex,time_to_mip,expected_value_solution=False,print_optimizer = False)
        results                 =   categorize_slots(input,results)
        #print_MSS(input, results) 
        saved_values            =   {}
        saved_values["input"]   =   input
        saved_values["results"] =   results
        pkl_name = 'first_stage_solution_'+ str(iter) + '_.pkl' 
        with open(pkl_name,"wb") as f:
            pickle.dump(saved_values,f)
        for n in  nScens:
            print("iter: %i  scen: %i "%(iter,n))
            col_index+=1 
            for seed in seeds:     
                """with open(pkl_name,"wb") as f:
                    pickle.dump(saved_values,f)
                input                   =   saved_values["input"]
                results                 =   saved_values["results"]"""
                results_i               =   copy.deepcopy(results)
                input                   =   generate_scenarios(input, n, seed) 
                results_i               =   run_greedy_construction_heuristic(input, results_i,debug=False)
                cell = alphabeth[col_index]+str(current_row_outlay + seed)
                sheet[cell] = results_i["obj"]
                #write_to_excel(excel_file_name,results_i,flex,n,seed)
        col_index=0
        current_row_outlay+= len(seeds) + 5
book.save("stability_test_matrix.xlsx")            
