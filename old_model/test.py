from openpyxl import Workbook
from openpyxl import load_workbook
def write_to_excel(file_name,input_dict,output_dict):
    #filename = "myfile.xlsx"
    new_row = []

    try:
        wb = load_workbook(file_name)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        headers_row = ['groups','scenarios','seed','flexibility','primal_bound', 'dual_bound','','MC',]
        for d in input_dict["Di"]:
            headers_row.append(d)
        headers_row.append('')
        headers_row.append('IC')
        for d in input_dict["Di"]:
            headers_row.append(d)
        wb = Workbook()
        ws = wb.active
        ws.append(headers_row)

    new_row.append(input_dict["number_of_groups"])
    new_row.append(input_dict["nScenarios"])
    new_row.append(input_dict["seed"])
    new_row.append(input_dict["F"])
    new_row.append(output_dict["obj"])
    new_row.append(output_dict["best_bound"])
    new_row.append('')
    new_row.append('')
    """for d in input_dict["Di"]:
        append"""

    ws.append(new_row)
    wb.save(file_name)