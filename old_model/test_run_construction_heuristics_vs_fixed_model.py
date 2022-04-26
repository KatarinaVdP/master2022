from functions_input import *
from functions_output import *
#from model_cutting_stock import *
from model_mip import *
from heuristic_greedy_construction import *
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import copy
import os.path


def run_model_mip_fixed2(input_dict,output_dict, time_limit, print_optimizer = False): 
    start_time =  time.time()
    #----- Sets ----- #  
    nDays           =   input_dict["nDays"]
    Wi  =   input_dict["Wi"]
    Si  =   input_dict["Si"]
    Gi  =   input_dict["Gi"]
    GWi =   input_dict["GWi"]
    GSi =   input_dict["GSi"]    
    Ri  =   input_dict["Ri"]
    RSi =   input_dict["RSi"]
    RGi =   input_dict["RGi"]
    Di  =   input_dict["Di"]
    Ci  =   input_dict["Ci"]

    #----- Parameter ----- #  
    F   =   input_dict["F"]
    E   =   input_dict["E"]
    TC  =   input_dict["TC"]
    I   =   input_dict["I"]
    B   =   input_dict["B"]
    H   =   input_dict["H"]
    K   =   input_dict["K"]
    L   =   input_dict["L"]
    U   =   input_dict["U"]
    N   =   input_dict["N"]
    T   =   input_dict["T"]
    Co  =   input_dict["Co"]
    J   =   input_dict["J"]
    P   =   input_dict["P"]
    Pi  =   input_dict["Pi"]
    Q   =   input_dict["Q"]
    Y   =   input_dict["Y"]
    nFixed = int(np.ceil((1-F) * sum(N[d] for d in Di)/I)*I)
    input_dict["nFixed"] = nFixed
    #----- Model ----- #
    m = gp.Model("mss_mip")
    m.setParam("TimeLimit", time_limit)
    #m.setParam("MIPFocus", 3) 
    # finding feasible solutions quickly:1
    # no trouble finding good quality solutions, more attention on proving optimality: 2 
    # If the best objective bound is moving very slowly (or not at all)and want to focus on the bound:3
    if not print_optimizer:
        m.Params.LogToConsole = 0
    
    '--- Variables ---'
    gamm    =   m.addVars(Si, Ri, Di, vtype=GRB.BINARY, name="gamma")
    lamb    =   m.addVars(Si, Ri, Di, vtype=GRB.BINARY, name="lambda")
    delt    =   m.addVars(Si, Ri, Di, Ci, vtype=GRB.BINARY, name="delta")
    x       =   m.addVars(Gi, Ri, Di, Ci, vtype=GRB.INTEGER, name="x")
    a       =   m.addVars(Gi, Ci, vtype=GRB.INTEGER, name="a")
    
    """#---- create from chategorize slot ----#
    for r in Ri:
        for d in Di:
            for s in Si:
                if output_dict["specialty_in_slot"][r][d]==s:
                    gamm[s,r,d].lb=1
                    gamm[s,r,d].ub=1
                else:
                    gamm[s,r,d].lb=0
                    gamm[s,r,d].ub=0
                if output_dict["extSlot"][r][d]==1:
                    lamb[s,r,d].lb=1
                    lamb[s,r,d].ub=1
                else:
                    lamb[s,r,d].lb=0
                    lamb[s,r,d].ub=0
    for g in Gi:
        for r in (list(set(Ri)^set(RGi[g]))):
            for d in Di:
                for c in Ci:  
                    x[g,r,d,c].lb=0
                    x[g,r,d,c].ub=0 """
    for s in Si:
        for r in Ri:
            for d in Di:
                gamm[s,r,d].lb=output_dict["gamm"][s][r][d]
                gamm[s,r,d].ub=output_dict["gamm"][s][r][d]
                lamb[s,r,d].lb=output_dict["lamb"][s][r][d]
                lamb[s,r,d].ub=output_dict["lamb"][s][r][d]
    for g in Gi:
        for r in (list(set(Ri)^set(RGi[g]))):
            for d in Di:
                for c in Ci:  
                    x[g,r,d,c].lb=0
                    x[g,r,d,c].ub=0
                    
    '--- Objective ---' 
    m.setObjective(
                quicksum(Pi[c] * Co[g] * a[g,c] for g in Gi for c in Ci)
    )   
    m.ModelSense = GRB.MINIMIZE 
    '--- Constraints ---'
    m.addConstr(
        quicksum( quicksum(gamm[s,r,d] for r in RSi[s]) for s in Si for d in Di) ==  nFixed ,
        name = "Con_PercentFixedRooms"
        )
    m.addConstrs(
        (lamb[s,r,d] <= gamm[s,r,d] 
        for s in Si for r in Ri for d in Di), 
        name = "Con_RollingFixedSlotCycles",
        )
    m.addConstrs(
        (quicksum(lamb[s,r,d] for r in RSi[s] for d in Di) <= U[s] 
        for s in Si),
        name = "Con_LongDaysCap",
    )
    print('Creating model (1/3)')
    m.addConstrs(
        (quicksum(gamm[s,r,d]+delt[s,r,d,c] for s in Si)<= 1 
        for r in Ri for d in Di for c in Ci),
        name= "Con_NoRoomDoubleBooking",
    )
    m.addConstrs(
        (quicksum(gamm[s,r,d]+delt[s,r,d,c] for r in RSi[s]) <= K[s][d] 
        for s in Si for d in Di for c in Ci),
        name= "Con_NoTeamDoubleBooking",
    )
    m.addConstrs(
        (quicksum(gamm[s,r,d]+delt[s,r,d,c] for s in Si for r in Ri) <= N[d] 
        for d in Di for c in Ci),
        name= "Con_TotalRoomsInUse",
    )
    for s in Si:
        m.addConstrs(
            (quicksum((L[g]+TC) * x[g,r,d,c] for g in GSi[s]) <= H[d] * (gamm[s,r,d] + delt[s,r,d,c]) + E*lamb[s,r,d] 
            for r in RSi[s] for d in Di for c in Ci),
        name = "Con_AvalibleTimeInRoom" + str(s), 
        )   
    m.addConstrs(
        (quicksum(x[g,r,d,c] for r in Ri for d in Di) + a[g,c] ==  Q[g][c] 
        for g in Gi for c in Ci),
        name= "Con_Demand",
    )
    m.addConstrs(
        (quicksum(delt[s,r,d,c] for s in Si) <= quicksum(x[g,r,d,c] for g in Gi) 
        for r in Ri for d in Di for c in Ci),
        name= "Con_OnlyAssignIfNecessary",
    )
    print('Creating model (2/3)')
    m.addConstrs(
        (quicksum(P[w][g][d-dd] * x[g,r,dd,c] for g in GWi[w] for r in Ri for dd in range(max(0,d+1-J[w]),d+1)) <= B[w][d] - Y[w][d] 
        for w in Wi for d in Di for c in Ci),
    name = "Con_BedOccupationCapacity",
    )
    for s in Si:
        m.addConstrs(
            (gamm[s,r,d]==gamm[s,r,nDays/I+d] 
            for r in RSi[s] for d in range(0,int(nDays-nDays/I))),
        name = "Con_RollingFixedSlotCycles" + str(s),
        )
    for s in Si:   
        m.addConstrs(
            (lamb[s,r,d]==lamb[s,r,nDays/I+d]  
            for r in RSi[s] for d in range(0,int(nDays-nDays/I))),
        name = "Con_RollingExtendedSlotCycles" + str(s),
        )
    print('Creating model (3/3)')
    create_model_time = (time.time() - start_time)
    print('Time to create model: %.1f s' % (create_model_time))    
    m.optimize()
    result_dict = save_results_pre(m)

    nSolutions=m.SolCount
    if nSolutions==0:
        result_dict["status"]=0
        print('Did not find any feasible initial solution in run_model_mip_fixed()')
        return
    else:
        #m.write('model.mps')      turn of during secons_stage_heuristic
        #m.write('warmstart.mst')               
        result_dict =  save_results(m, input_dict, result_dict)
    return result_dict

def print_heuristic_vs_fixed(results_mip: dict, results_heuristic: dict, flex: int, nScenarios: int, seed: int,max_time_fixed_mip: int):
    
    mip_str_p = "mip_p: "+ str("{:.1f}".format(results_mip["obj"]))
    mip_str_d = "mip_d: "+ str("{:.1f}".format(results_mip["best_bound"]))
    heur_str =  "heur: "  + str("{:.1f}".format(results_heuristic["obj"]))
    preformance_str = "diff: " + str("{:.2f}".format((results_mip["obj"]-results_heuristic["obj"])/results_mip["obj"])) 
    

    print("{0:<15}".format(flex), end="")
    print("{0:<15}".format(nScenarios), end="")
    print("{0:<15}".format(seed), end="")
    print("{0:<15}".format("mip_time: " + str(max_time_fixed_mip)), end="")
    print("{0:<15}".format(mip_str_p), end="")
    print("{0:<15}".format(mip_str_d),end="")
    print("{0:<15}".format(heur_str),end="")
    print("{0:<15}".format(preformance_str),end="")
    print() 

def write_to_excel(excel_file_name: str, results_mip: dict, results_heuristic: dict, flex: int, nScenarios: int, seed: int,max_time_fixed_mip: int):
    #filename = "myfile.xlsx"

    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]# select first worksheet
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("Heur_vs_mip",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("max_time_fixed_mip")
        header_row.append("primal")
        header_row.append("dual")
        header_row.append("MIPGap")
        header_row.append("heuristic")
        header_row.append("heuristic_time")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(max_time_fixed_mip)
    new_row.append(results_mip["obj"])
    new_row.append(results_mip["best_bound"])
    new_row.append(results_mip["MIPGap"])
    new_row.append(results_heuristic["obj"])
    new_row.append(results_heuristic["heuristic_time"])
    ws.append(new_row)
    wb.save(excel_file_name)  

def write_to_excel_all(excel_file_name: str, flex: int, nScenarios: int, seed: int,max_time_fixed_mip: int, results_mip: dict, results_to_print: list):
    #filename = "myfile.xlsx"

    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]# select first worksheet
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("Heur_vs_mip",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("max_time_fixed_mip")
        header_row.append("primal")
        header_row.append("dual")
        header_row.append("MIPGap")
        header_row.append("heuristic")
        header_row.append("time")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(max_time_fixed_mip)
    new_row.append(results_mip["obj"])
    new_row.append(results_mip["best_bound"])
    new_row.append(results_mip["MIPGap"])
    for i in range(len(results_to_print)):
        new_row.append(results_to_print[i])
    ws.append(new_row)
    wb.save(excel_file_name)      
print('initializing...')
number_of_groups            =   25
MC_cap                      =   [60,49]         #[weekday,weekend]                          #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
IC_cap                      =   [11,6] #[weekday,weekend]      
"""MC_cap                      =   [30,24.5]         #[weekday,weekend]                          #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
IC_cap                      =   [5.5,3] """
nScenarios                  =   500
"""flexibilities               =   [0,0.05,0.1,0.15,0.20,0.25,0.30,0.35,0.40,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]"""
flexibilities               =   [0.15,0]
num_sols_to_investigate     =   3   
seeds                       =   [i for i in range(1,num_sols_to_investigate+1)]

time_to_mip                 =   30
nScenarios_initial_sol      =   3
max_time_fixed_mip          =   300
file_name                   =   choose_correct_input_file(number_of_groups)
excel_file_name             =   'input_output/test_heuristics_'+ str(number_of_groups) + '_groups_10_scen.xlsx'
for flex in flexibilities:
    input                   =   read_input(file_name)                 
    input                   =   change_ward_capacity(input, "MC", MC_cap[0],  MC_cap[1])
    input                   =   change_ward_capacity(input, "IC", IC_cap[0],  IC_cap[1])
    if number_of_groups ==   25:
        change_demand(input, 1.35, print_minutes = False)
    for seed in seeds:
        
        input                   =   generate_scenarios(input, nScenarios_initial_sol, seed)
        results, input          =   run_model_mip(input,flex,time_to_mip,expected_value_solution=False,print_optimizer = False)
        results                 =   categorize_slots(input,results)
        print_MSS(input, results) 

        input                   =   generate_scenarios(input, nScenarios, seed)
        
        results_m                           =   copy.deepcopy(results)   
        results_h                           =   copy.deepcopy(results)
        results_h_smart_flex                =   copy.deepcopy(results)
        results_h_smarter_flex              =   copy.deepcopy(results)
        results_h_smart_fix                 =   copy.deepcopy(results)
        results_h_smart_fix_smart_flex      =   copy.deepcopy(results)
        results_h_smart_fix_smarter_flex    =   copy.deepcopy(results)
        results_h_smarter_fix               =   copy.deepcopy(results)
        results_h_smarter_fix_smart_flex    =   copy.deepcopy(results)
        results_h_smarter_fix_smarter_flex  =   copy.deepcopy(results)
        
        results_m                           =   run_model_mip_fixed2(input,results_m,max_time_fixed_mip,print_optimizer = True)
        results_h                           =   run_greedy_construction_heuristic(input,results_h)
        results_h_smart_flex                =   run_greedy_construction_heuristic_smart_flex(input,results_h_smart_flex)
        results_h_smarter_flex              =   run_greedy_construction_heuristic(input,results_h_smarter_flex)
        results_h_smart_fix                 =   run_greedy_construction_heuristic_smart_fix(input,results_h_smart_fix)
        results_h_smart_fix_smart_flex      =   run_greedy_construction_heuristic_smart_fix_smart_flex(input,results_h_smart_fix_smart_flex)
        results_h_smart_fix_smarter_flex    =   run_greedy_construction_heuristic_smart_fix_smarter_flex(input,results_h_smart_fix_smarter_flex)
        results_h_smarter_fix               =   run_greedy_construction_heuristic_smarter_fix(input,results_h_smarter_fix)
        results_h_smarter_fix_smart_flex    =   run_greedy_construction_heuristic_smarter_fix_smart_flex(input,results_h_smarter_fix_smart_flex)
        results_h_smarter_fix_smarter_flex  =   run_greedy_construction_heuristic_smarter_fix_smarter_flex(input,results_h_smarter_fix_smarter_flex)
    
        print("flex: %.2f  seed: %i  MIPgap: %.3f "%(flex,seed, results_m["MIPGap"]))
        """print("best bound fixed MIP:            " + "{0:<10}".format("{:.1f}".format(results_m["obj"])) )
        print("heur:                            " + "{0:<10}".format("{:.1f}".format(results_h["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h["heuristic_time"])))
        print("heur_smart_flex:                 " + "{0:<10}".format("{:.1f}".format(results_h_smart_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smart_flex["heuristic_time"])))
        print("heur_smarter_flex:               " + "{0:<10}".format("{:.1f}".format(results_h_smarter_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smarter_flex["heuristic_time"])))
        print("heur_smart_fix:                  " + "{0:<10}".format("{:.1f}".format(results_h_smart_fix["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smart_fix["heuristic_time"])))
        print("heur_smart_fix_smart_flex:       " + "{0:<10}".format("{:.1f}".format(results_h_smart_fix_smart_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smart_fix_smart_flex["heuristic_time"])))
        print("heur_smart_fix_smarter_flex:     " + "{0:<10}".format("{:.1f}".format(results_h_smart_fix_smarter_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smart_fix_smarter_flex["heuristic_time"])))
        print("heur_smarter_fix:                " + "{0:<10}".format("{:.1f}".format(results_h_smarter_fix["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smarter_fix["heuristic_time"])))
        print("heur_smarter_fix_smart_flex:     " + "{0:<10}".format("{:.1f}".format(results_h_smarter_fix_smart_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smarter_fix_smart_flex["heuristic_time"])))
        print("heur_smarter_fix_smarter_flex:   " + "{0:<10}".format("{:.1f}".format(results_h_smarter_fix_smarter_flex["obj"])) + "time: " + "{0:<5}".format("{:.2f}".format(results_h_smarter_fix_smarter_flex["heuristic_time"])))"""
        
        #results_heuristic       =   run_greedy_construction_heuristic_smart_flex(input, results_heuristic,debug=False)
        #results_mip   =   run_model_mip_fixed2(input,results_mip,max_time_fixed_mip)
        
        results_to_print = [
            results_h["obj"],                           results_h["heuristic_time"],
            results_h_smart_flex["obj"],                results_h_smart_flex["heuristic_time"],
            results_h_smarter_flex["obj"],              results_h_smarter_flex["heuristic_time"],
            results_h_smart_fix["obj"],                 results_h_smart_fix["heuristic_time"],
            results_h_smart_fix_smart_flex["obj"],      results_h_smart_fix_smart_flex["heuristic_time"],
            results_h_smart_fix_smarter_flex["obj"],    results_h_smart_fix_smarter_flex["heuristic_time"],
            results_h_smarter_fix["obj"],               results_h_smarter_fix["heuristic_time"],
            results_h_smarter_fix_smart_flex["obj"],    results_h_smarter_fix_smart_flex["heuristic_time"],
            results_h_smarter_fix_smarter_flex["obj"],  results_h_smarter_fix_smarter_flex["heuristic_time"],
        ]
        write_to_excel_all(excel_file_name, flex, nScenarios, seed,max_time_fixed_mip, results_m, results_to_print)
        #print_heuristic_vs_fixed(results_mip,results_heuristic, flex, nScenarios,seed,max_time_fixed_mip)
        
