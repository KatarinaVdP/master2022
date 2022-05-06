from functions_input import *
from functions_output import *
#from model_cutting_stock import *
from model_mip import *
from heuristic_greedy_construction import *
import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import copy
import os.path

def write_to_excel(excel_file_name: str, results_heuristic: dict, flex: int, nScenarios: int, seed: int):
    #filename = "myfile.xlsx"

    try:
        wb = load_workbook(excel_file_name)
        ws = wb.worksheets[0]# select first worksheet
    except FileNotFoundError:
        wb  = Workbook()
        ws = wb.create_sheet("stability analysis",0)
        wb.save(excel_file_name)
        ws = wb.worksheets[0]
        header_row=[]
        header_row.append("flex")
        header_row.append("nScenarios")
        header_row.append("seed")
        header_row.append("heuristic_sol")
        header_row.append("heuristic_time")
        ws.append(header_row)
        wb.save(excel_file_name) 
    new_row = []  
    new_row.append(flex)
    new_row.append(nScenarios)
    new_row.append(seed)
    new_row.append(results_heuristic["obj"])
    new_row.append(results_heuristic["heuristic_time"])
    ws.append(new_row)
    wb.save(excel_file_name)  

print('initializing...')
#---- change parameters ----
'---number og groups--- 9/25'
number_of_groups            =   25                                                          #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
'---files---'
file_name                   =   choose_correct_input_file(number_of_groups)
excel_file_name             =   'input_output/stability_test_matrix_'+ str(number_of_groups) + '_groups_new.xlsx'
'---input parameter tuning---'
input                       =   read_input(file_name)

'---loops---'
flexibilities               =   [0]
num_sols_to_investigate     =   30                                                           #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
num_seeds_per_solution      =   30                                                          #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
start_nScenarios            =   10                                                          #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
end_nScenarios              =   600                                                         #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
gap_nScenarios              =   10
nScens                      =   [i*gap_nScenarios for i in range(int(start_nScenarios/gap_nScenarios),int(end_nScenarios/gap_nScenarios) + 1)]
seeds                       =   [i for i in range(1,num_seeds_per_solution+1)]
'---initial solution generation---'
nScenarios_initial_sol      =   1
time_to_mip                 =   30
mip_seeds                   =   [i for i in range(101,101 + num_seeds_per_solution)]


#---- Excel Layout parameters ----
book                        =   Workbook()
sheet                       =   book.active
current_row_outlay          =   2
col_index                   =   0
alphabeth   =   ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M', 'N', 'O', 'P', 'Q','R', 'S','T','U','V','W','X','Y','Z',
                'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK','AL','AM', 'AN', 'AO', 'AP', 'AQ','AR', 'AS','AT','AU','AV','AW','AX','AY','AZ',
                'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK','BL','BM', 'BN', 'BO', 'BP', 'BQ','BR', 'BS','BT','BU','BV','BW','BX','BY','BZ',
                'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK','CL','CM', 'CN', 'CO', 'CP', 'CQ','CR', 'CS','CT','CU','CV','CW','CX','CY','CZ']

for flex in flexibilities:
    for iter in range(num_sols_to_investigate):
        pkl_name = 'first_stage_solution_'+ str(mip_seeds[iter]) + '_.pkl' 
        try:
            with open(pkl_name,"rb") as f:
                saved_values        =   pickle.load(f)
                print('Reading first stage solution from pkl')
                input               =   saved_values["input"]
                results             =   saved_values["results"]
        except FileNotFoundError:
            print('Grenerating new first stage solution')
            input                   =   generate_scenarios(input, nScenarios_initial_sol, mip_seeds[iter])
            results, input          =   run_model_mip(input,flex,time_to_mip,expected_value_solution=False,print_optimizer = False)
            results                 =   categorize_slots(input,results)
            saved_values            =   {}
            saved_values["input"]   =   input
            saved_values["results"] =   results
            with open(pkl_name,"wb") as f:
                pickle.dump(saved_values,f)
        for n in  nScens:
            print("iter: %i  scen: %i "%(iter,n))
            col_index+=1 
            for seed in seeds:     
                results_i           =   copy.deepcopy(results)
                input               =   generate_scenarios(input, n, seed) 
                results_i           =   run_greedy_construction_heuristic(input, results_i,debug=False)
                cell                =   alphabeth[col_index]+str(current_row_outlay + seed)
                sheet[cell]         =   results_i["obj"]
        col_index                   =   0
        current_row_outlay          +=  len(seeds) + 5
        book.save(excel_file_name)            
